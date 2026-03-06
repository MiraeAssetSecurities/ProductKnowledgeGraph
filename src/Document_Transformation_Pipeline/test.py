import base64
import io
import json
import logging
import os
import requests
import shutil
import subprocess
import tempfile
import textwrap
from pathlib import Path
from typing import Iterable, Mapping, Sequence

from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from docling.datamodel.base_models import InputFormat
from docling.datamodel.pipeline_options import VlmPipelineOptions
from docling.datamodel.pipeline_options_vlm_model import ApiVlmOptions, ResponseFormat
from docling.document_converter import (
    DocumentConverter,
    PdfFormatOption,
)
from docling.pipeline.vlm_pipeline import VlmPipeline

try:
    import pypdfium2
    PDF_RENDER_AVAILABLE = True
except Exception:  # pragma: no cover
    PDF_RENDER_AVAILABLE = False


def openai_compatible_vlm_options(
    model: str,
    prompt: str,
    format: ResponseFormat,
    hostname_and_port: str,
    temperature: float = 0.2,
    max_tokens: int = 4096,
    api_key: str = "",
    skip_special_tokens: bool = False,
):
    headers = {}
    if api_key:
        headers["Authorization"] = f"Bearer {api_key}"

    return ApiVlmOptions(
        url=f"http://{hostname_and_port}/v1/chat/completions",
        params=dict(
            model=model,
            max_tokens=max_tokens,
            #skip_special_tokens=skip_special_tokens,
        ),
        headers=headers,
        prompt=prompt,
        timeout=90,
        scale=2.0,
        temperature=temperature,
        response_format=format,
    )

def build_vlm_pipeline_options() -> VlmPipelineOptions:
    options = VlmPipelineOptions(enable_remote_services=True)
    default_prompt = textwrap.dedent(
        """
        너는 복잡한 문서의 레이아웃을 분석하고 마크다운으로 변환하는 전문가야.
        내부 시스템 구조도나 해석하기 어려운 시각 요소는 억지로 텍스트화하지 말고, 아래 지침에 따라 처리해줘.

        ### 1. 문서 개요 (Document Summary)
        - 문서의 최상단에 다음과 같은 형식으로 요약을 작성한다.
        - **문서 유형:** [예: 내부 시스템 아키텍처 가이드, 운영 매뉴얼, 비즈니스 보고서 등]
        - **핵심 목적:** [이 문서가 전달하고자 하는 핵심 내용을 1~2문장으로 요약]

        ### 2. 본문 및 레이아웃 OCR
        - 텍스트는 마크다운(Markdown) 문법을 준수하여 추출한다.
        - 제목(H1, H2, H3)과 리스트(-, 1.) 구조를 원본과 동일하게 유지한다.
        - 표(Table)는 행/열 구조를 유지하여 마크다운 표 형식으로 변환한다.

        ### 3. 이미지 및 시스템 도식 처리 (Image Handling)
        - 해석이 불가능한 복잡한 다이어그램, 사진, 혹은 내부 시스템 화면 캡처는 다음과 같이 처리한다.
        - **태그 삽입:** 이미지가 위치한 자리에 [IMAGE_PLACEHOLDER: {index}] 태그를 남긴다. (예: [IMAGE_PLACEHOLDER: 1])
        - **시각적 묘사:** 태그 바로 아래에 > [이미지 설명]:  블록을 만들고, 해당 이미지가 무엇을 의미하는지(예: 로그인 프로세스 흐름도, 서버 구성도 등) 상세히 기술한다. 이는 RAG의 검색 성능을 높이기 위함이다.

        ### 4. RAG 및 청킹 최적화
        - 섹션이 바뀔 때마다 --- (Horizontal Rule)를 사용하여 명확히 구분한다.
        - 페이지 번호나 반복되는 머리말/꼬리말은 삭제하여 문맥의 연속성을 확보한다.

        ---
        [출력 예시]
        ## Document Overview
        - **문서 유형:** 시스템 구성 가이드
        - **핵심 목적:** 신규 입사자를 위한 내부 클라우드 인프라 구조 및 권한 신청 프로세스 설명

        ## 1. 인프라 개요
        본 시스템은 마이크로서비스 아키텍처를 따릅니다.

        [IMAGE_PLACEHOLDER: 1]
        > [이미지 설명]: VPC 내부에 위치한 쿠버네티스 클러스터와 외부 로드밸런서 간의 통신 구조를 나타내는 아키텍처 다이어그램임.
        """
    ).strip()
    options.vlm_options = openai_compatible_vlm_options(
        model=os.getenv("VLM_MODEL", "bedrock-qwen3-v1"),
        hostname_and_port=os.getenv("VLM_HOST", "54.197.26.233:4000"),
        prompt=os.getenv("VLM_PROMPT", default_prompt),
        format=ResponseFormat.MARKDOWN,
        api_key=os.getenv("VLM_API_KEY", "KEY_AAA"),
    )
    return options

def make_json_safe(obj):
    try:
        from pydantic import AnyUrl
    except Exception:  # pragma: no cover
        AnyUrl = ()

    if isinstance(obj, AnyUrl):
        return str(obj)
    if isinstance(obj, Mapping):
        return {k: make_json_safe(v) for k, v in obj.items()}
    if isinstance(obj, Sequence) and not isinstance(obj, (str, bytes)):
        return [make_json_safe(v) for v in obj]
    return obj

def save_outputs(markdown_text: str, json_obj, stem: str, out_dir: Path):
    out_dir.mkdir(parents=True, exist_ok=True)
    (out_dir / f"{stem}.md").write_text(markdown_text, encoding="utf-8")
    safe_lossless = make_json_safe(json_obj)
    with open(out_dir / f"{stem}.lossless.json", "w", encoding="utf-8") as f:
        json.dump(safe_lossless, f, ensure_ascii=False, indent=2)

def dataframe_like_to_markdown(rows):
    if not rows:
        return "(empty sheet)"
    widths = [max(len(str(v)) for v in col) for col in zip(*rows)]
    lines = []
    header = rows[0]
    header_line = "| " + " | ".join(str(v).ljust(widths[i]) for i, v in enumerate(header)) + " |"
    sep_line = "|" + "|".join("-" * (widths[i] + 2) for i in range(len(widths))) + "|"
    lines.append(header_line)
    lines.append(sep_line)
    for row in rows[1:]:
        lines.append("| " + " | ".join(str(v).ljust(widths[i]) for i, v in enumerate(row)) + " |")
    return "\n".join(lines)

def convert_xlsx_standard(xlsx_paths: Iterable[Path], out_dir: Path):
    results = {}
    for path in xlsx_paths:
        wb = load_workbook(path, data_only=True)
        md_parts = []
        json_data = {}
        for sheet in wb.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                md_parts.append(f"## {sheet.title}\n(empty sheet)\n")
                json_data[sheet.title] = {"table": [], "images": []}
                continue
            # ensure header row exists; if not, synthesize
            header = rows[0]
            if any(h is None for h in header):
                header = [f"col{i+1}" for i in range(len(header))]
                rows = [header] + rows
            markdown_table = dataframe_like_to_markdown([["" if v is None else v for v in r] for r in rows])
            md_parts.append(f"## {sheet.title}\n{markdown_table}\n")
            data_rows = []
            for row in rows[1:]:
                record = {str(header[i]): row[i] for i in range(len(header))}
                data_rows.append(record)
            images_info = []
            images_md = []
            for idx, img in enumerate(getattr(sheet, "_images", []), start=1):
                anchor = getattr(img, "anchor", None)
                cell_ref = None
                if hasattr(anchor, "_from"):
                    cell_ref = f"{get_column_letter(anchor._from.col + 1)}{anchor._from.row + 1}"
                mime = getattr(img, "_mimetype", "image/png")
                data_bytes = img._data() if hasattr(img, "_data") else None
                data_uri = None
                if data_bytes:
                    b64 = base64.b64encode(data_bytes).decode("ascii")
                    data_uri = f"data:{mime};base64,{b64}"
                    images_md.append(f"![{sheet.title}-img{idx} at {cell_ref or '?'}]({data_uri})")
                else:
                    images_md.append(f"(이미지 {idx}, 위치 {cell_ref or '?'}, 데이터 없음)")
                images_info.append(
                    {
                        "index": idx,
                        "cell": cell_ref,
                        "mime": mime,
                        "data_uri": data_uri,
                    }
                )
            if images_md:
                md_parts.append("### 이미지\n" + "\n\n".join(images_md))
            json_data[sheet.title] = {"table": data_rows, "images": images_info}
        markdown_text = "\n".join(md_parts)
        save_outputs(markdown_text, json_data, Path(path).stem, out_dir / "standard")
        logging.info("표준(XLSX→openpyxl) 변환 완료: %s", path.name)
        results[Path(path).stem] = {"markdown": markdown_text, "json": json_data}
    return results

def convert_pdf_with_vlm(pdf_paths: Iterable[Path], out_dir: Path):
    pipeline_options = build_vlm_pipeline_options()
    converter = DocumentConverter(
        format_options={
            InputFormat.PDF: PdfFormatOption(
                pipeline_cls=VlmPipeline,
                pipeline_options=pipeline_options,
            )
        }
    )
    results = {}
    for pdf_path in pdf_paths:
        result = converter.convert(pdf_path)
        stem = Path(pdf_path).stem
        doc = result.document
        markdown_text = doc.export_to_markdown()
        lossless = doc.export_to_dict(mode="lossless")
        save_outputs(markdown_text, lossless, stem, out_dir / "vlm")
        logging.info("VLM 파이프라인 변환 완료: %s", pdf_path.name)
        results[stem] = {"markdown": markdown_text, "json": lossless}
    return results

def convert_xlsx_to_pdf(xlsx_path: Path, out_dir: Path):
    if shutil.which("libreoffice") is None:
        logging.warning("LibreOffice가 없어 PDF 변환을 건너뜁니다: %s", xlsx_path.name)
        return None

    out_dir.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir_path = Path(tmpdir)
        cmd = [
            "libreoffice",
            "--headless",
            "--convert-to",
            'pdf:calc_pdf_Export:{"SinglePageSheets":{"type":"boolean","value":"true"},"FitWidthToPage":{"type":"boolean","value":"true"}}',
            str(xlsx_path),
            "--outdir",
            str(tmpdir_path),
        ]
        try:
            subprocess.run(cmd, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        except subprocess.CalledProcessError as exc:
            logging.error("LibreOffice 변환 실패: %s", exc)
            return None

        generated = next(tmpdir_path.glob("*.pdf"), None)
        if not generated:
            logging.error("PDF 변환 산출물이 없습니다: %s", xlsx_path.name)
            return None

        target = out_dir / generated.name
        shutil.move(str(generated), target)
        return target

def convert_xlsx_dual_pass(xlsx_paths: Iterable[Path], out_dir: Path):
    std_results = convert_xlsx_standard(xlsx_paths, out_dir)

    pdf_list = []
    for path in xlsx_paths:
        pdf_path = convert_xlsx_to_pdf(path, out_dir / "tmp_pdf")
        if pdf_path:
            pdf_list.append(pdf_path)
    vlm_results = convert_pdf_with_vlm(pdf_list, out_dir) if pdf_list else {}
    if not vlm_results:
        logging.warning("VLM용 PDF가 없어 VLM 변환을 건너뜁니다.")

    for path in xlsx_paths:
        stem = Path(path).stem
        std_res = std_results.get(stem)
        vlm_res = vlm_results.get(stem)

        if not std_res and not vlm_res:
            logging.warning("결합할 결과가 없습니다: %s", stem)
            continue

        # 마크다운 결합: 표준 테이블/텍스트를 기본으로 하고, VLM 해석/이미지 캡션을 추가 섹션으로 부착
        md_parts = []
        if std_res:
            md_parts.append("# 표준 파싱 결과\n")
            md_parts.append(std_res["markdown"])
        if vlm_res:
            md_parts.append("# VLM 보강 결과 (이미지/레이아웃 기반)\n")
            md_parts.append(vlm_res["markdown"])
        combined_md = "\n\n".join(md_parts)

        if os.getenv("PRESERVE_IMAGES", "0") == "1":
            logging.info("PRESERVE_IMAGES")
            pdf_path = out_dir / "tmp_pdf" / f"{stem}.pdf"
            img_md = pdf_to_base64_image_markdown(pdf_path)
            if img_md:
                combined_md = combined_md + "\n\n# 원본 이미지 스냅샷\n" + img_md

        if os.getenv("USE_LLM_MERGE", "0") == "1" and std_res and vlm_res:
            #logging.info("USE_LLM_MERGE")
            merged = merge_markdown_with_llm(std_res["markdown"], vlm_res["markdown"])
            if merged:
                combined_md = merged

        combined_json = {}
        if vlm_res:
            combined_json["vlm_lossless"] = vlm_res["json"]
        if std_res:
            combined_json["standard_tables"] = std_res["json"]

        if vlm_res:
            combined_json["main"] = vlm_res["json"]
        elif std_res:
            combined_json["main"] = std_res["json"]

        save_outputs(combined_md, combined_json, stem, out_dir / "combined")
        logging.info("표준+VLM 결합 저장 완료: %s", stem)

def merge_markdown_with_llm(std_md: str, vlm_md: str) -> str | None:
    """
    LLM에 표준/스캔 결과를 함께 주고 통합된 단일 마크다운을 생성한다.
    - 표준 테이블/텍스트를 우선 사용
    - VLM이 제공하는 이미지 캡션/비정형 해석을 필요 부분에 녹여 요약
    반환 실패 시 None
    """
    host = os.getenv("VLM_HOST", "54.197.26.233:4000")
    model = os.getenv("VLM_MODEL", "bedrock-qwen3-v1")
    api_key = os.getenv("VLM_API_KEY", "KEY_AAA")
    url = f"http://{host}/v1/chat/completions"

    prompt = textwrap.dedent(
        f"""
        너는 엑셀 파싱 결과를 정리하는 마크다운 편집 전문가야.
        standard_md의 텍스트 데이터를 vlm_md의 마크다운 레이아웃에 넣고, vlm_md에 포함된 부가 설명을 첨부해서 마크다운 결과물을 생성하는 거야.
        
        ### 핵심 원칙 (Core Principles)
        1. 레이아웃은 vlm_md 기준: 전체적인 문서의 흐름, 마크다운 형식(표의 형태, 섹션 구분, 리스트 스타일 등)은 시각적으로 깔끔한 vlm_md의 스타일을 그대로 따른다.
        2. 데이터는 standard_md 기준: standard_md에 포함된 텍스트 데이터는 반드시 포함시킨다. 
        3. 이미지 보존: standard_md에 없는 이미지나 차트는 vlm_md의 상세 설명을 활용하여 위치시킨다.

        ### 세부 지침 (Instructions)

        1. 문서 구조화:
        - # Sheet: [시트명]으로 대분류를 시작한다.
        - vlm_md가 보여주는 문서의 시각적 위계(H2, H3)를 유지하여 가독성을 높인다.

        2. 표(Table) 재구성:
        - std_md의 col1, col2 같은 표현은 사용하지 않는다.
        - vlm_md에서 파악한 표의 제목과 열 이름을 사용하되, 그 안의 값은 std_md의 데이터를 채워 넣는다.
        - 엑셀의 병합된 셀이나 복잡한 구조는 마크다운 표 표준 규격에 맞춰 자연스럽게 변환한다.
        - 표가 필요한 경우에는 표 형식으로 표현한다.

        3. 시각적 요소 및 OCR 처리:
        - 이미지나 도형은 vlm_md의 설명을 따르며, 해석이 불가능한 시스템 UI 등은 ![설명](base64_data_placeholder) 형태로 유지한다.
        - std_md에 없는 시각적 정보는 반드시 인용구(>) 등을 활용해 본문과 구분하여 보충한다.

        4. RAG 최적화:
        - 각 청크(Chunk)가 분리되어도 맥락을 알 수 있게 상단에 [시트명 > 섹션명] 정보를 포함한다.
        - 단순 나열된 데이터는 검색이 용이하도록 서술형 문장을 살짝 덧붙인다.

        ### 출력 형식 (Output Format)
        - Markdown 포맷만 출력할 것.
        - 시각적으로 유려한 마크다운 형식을 취하되, 데이터는 엑셀 원본(std_md)의 무결성을 유지할 것.

        ---
        [standard_md]
        {std_md}

        [vlm_md]
        {vlm_md}
        """
    )

    headers = {"Content-Type": "application/json"}
    if api_key:
        headers["Authorization"] = f"Bearer {api_key}"

    body = {
        "model": model,
        "messages": [
            {
                "role": "system",
                "content": "You are a helpful assistant that merges markdown outputs from structured (standard) and OCR (VLM) sources.",
            },
            {"role": "user", "content": prompt},
        ],
        "max_tokens": 4096,
        "temperature": 0.3,
    }

    try:
        resp = requests.post(url, headers=headers, json=body, timeout=90)
        resp.raise_for_status()
        data = resp.json()
        return data["choices"][0]["message"]["content"]
    except Exception as exc:
        logging.error("LLM 결합 실패: %s", exc)
        return None

def pdf_to_base64_image_markdown(pdf_path: Path) -> str | None:
    if not PDF_RENDER_AVAILABLE:
        logging.info("pypdfium2가 없어 원본 이미지 스냅샷을 건너뜁니다.")
        return None
    if not pdf_path.exists():
        return None
    try:
        pdf = pypdfium2.PdfDocument(str(pdf_path))
        md_parts = []
        for i in range(len(pdf)):
            page = pdf[i]
            bitmap = page.render(scale=1.5)
            pil_image = bitmap.to_pil()
            buf = io.BytesIO()
            pil_image.save(buf, format="PNG")
            b64 = base64.b64encode(buf.getvalue()).decode("ascii")
            md_parts.append(f"![page {i+1}](data:image/png;base64,{b64})")
        return "\n\n".join(md_parts)
    except Exception as exc:
        logging.error("PDF→이미지 변환 실패: %s", exc)
        return None

def main():
    logging.basicConfig(level=logging.INFO)

    input_paths = [
        Path("[8951] 표준메시지_반대의사_20210305.xlsx"),
        #Path("법인 고객유형코드 및 투자자분류코드 정리_20250731.xlsx"),
    ]
    out_dir = Path("output_dual")
    convert_xlsx_dual_pass(input_paths, out_dir)
    
if __name__ == "__main__":
    main()